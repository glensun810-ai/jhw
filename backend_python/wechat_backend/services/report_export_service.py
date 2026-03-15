"""
诊断报告导出服务

功能:
- 导出 Excel 格式报告
- 导出 HTML 格式报告（用于转长图）
- 支持批量导出

@author: 系统架构组
@date: 2026-03-14
@version: 1.0.0
"""

import io
import json
from typing import Dict, Any, List, Optional
from datetime import datetime
from pathlib import Path

from wechat_backend.logging_config import api_logger

# 尝试导入 openpyxl，如果不存在则返回 None
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    api_logger.warning("[ExportService] openpyxl 未安装，Excel 导出功能不可用")


class ReportExportService:
    """
    报告导出服务（P1-02 实现版 - 2026-03-14）
    
    功能:
    1. 导出 Excel 格式报告
    2. 导出 HTML 格式报告（用于转长图）
    3. 支持批量导出
    """
    
    def __init__(self):
        """初始化导出服务"""
        # Excel 样式配置
        self.header_font = Font(bold=True, size=12, color='FFFFFF')
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        self.cell_font = Font(size=11)
        self.cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 边框样式
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_to_excel(self, report_data: Dict[str, Any], execution_id: str = None) -> bytes:
        """
        导出报告为 Excel 格式
        
        参数:
            report_data: 报告数据
            execution_id: 执行 ID（用于文件名）
            
        返回:
            bytes: Excel 文件二进制数据
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl 未安装，请运行：pip install openpyxl")
        
        try:
            # 创建工作簿
            wb = Workbook()
            
            # 1. 概况表
            ws_summary = wb.active
            ws_summary.title = "诊断概况"
            self._create_summary_sheet(ws_summary, report_data)
            
            # 2. 品牌分布表
            ws_brand = wb.create_sheet(title="品牌分布")
            self._create_brand_distribution_sheet(ws_brand, report_data)
            
            # 3. 情感分析表
            ws_sentiment = wb.create_sheet(title="情感分析")
            self._create_sentiment_sheet(ws_sentiment, report_data)
            
            # 4. 关键词表
            ws_keywords = wb.create_sheet(title="关键词")
            self._create_keywords_sheet(ws_keywords, report_data)
            
            # 5. 详细结果表
            ws_results = wb.create_sheet(title="详细结果")
            self._create_results_sheet(ws_results, report_data)
            
            # 调整列宽
            self._auto_adjust_column_width(wb)
            
            # 保存到内存
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            api_logger.info(f"[ExportService] ✅ Excel 导出成功：{execution_id or 'N/A'}")
            
            return output.getvalue()
            
        except Exception as e:
            api_logger.error(f"[ExportService] ❌ Excel 导出失败：{e}", exc_info=True)
            raise
    
    def _create_summary_sheet(self, ws, report_data: Dict[str, Any]):
        """创建概况表"""
        # 标题
        ws['A1'] = 'AI 品牌诊断报告'
        ws['A1'].font = Font(bold=True, size=18)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # 基本信息
        basic_info = [
            ('报告 ID', report_data.get('execution_id', 'N/A')),
            ('品牌名称', report_data.get('brand_name', 'N/A')),
            ('诊断状态', report_data.get('status', 'N/A')),
            ('创建时间', report_data.get('created_at', 'N/A')),
            ('完成时间', report_data.get('completed_at', 'N/A')),
            ('结果数量', len(report_data.get('results', []))),
        ]
        
        row = 3
        for label, value in basic_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # 竞品列表
        competitors = report_data.get('competitor_brands', [])
        if competitors:
            ws[f'A{row+1}'] = '竞品列表'
            ws[f'A{row+1}'].font = Font(bold=True)
            for i, comp in enumerate(competitors, row+2):
                ws[f'A{i}'] = f'  • {comp}'
    
    def _create_brand_distribution_sheet(self, ws, report_data: Dict[str, Any]):
        """创建品牌分布表"""
        # 表头
        headers = ['品牌名称', '提及次数', '占比']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        # 数据
        brand_dist = report_data.get('brandDistribution', {})
        data = brand_dist.get('data', {})
        total_count = brand_dist.get('totalCount', 0)
        
        row = 2
        for brand, count in (data or {}).items():
            percentage = f'{(count / total_count * 100):.1f}%' if total_count > 0 else '0%'
            
            ws.cell(row=row, column=1, value=brand).border = self.thin_border
            ws.cell(row=row, column=2, value=count).border = self.thin_border
            ws.cell(row=row, column=3, value=percentage).border = self.thin_border
            row += 1
        
        # 总计
        ws.cell(row=row, column=1, value='总计').font = Font(bold=True)
        ws.cell(row=row, column=2, value=total_count).font = Font(bold=True)
        ws.cell(row=row, column=3, value='100%').font = Font(bold=True)
    
    def _create_sentiment_sheet(self, ws, report_data: Dict[str, Any]):
        """创建情感分析表"""
        # 表头
        headers = ['情感类型', '数量', '占比']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        # 数据
        sentiment_dist = report_data.get('sentimentDistribution', {})
        data = sentiment_dist.get('data', {})
        total_count = sentiment_dist.get('totalCount', 0)
        
        row = 2
        for sentiment, count in (data or {}).items():
            percentage = f'{(count / total_count * 100):.1f}%' if total_count > 0 else '0%'
            
            ws.cell(row=row, column=1, value=sentiment).border = self.thin_border
            ws.cell(row=row, column=2, value=count).border = self.thin_border
            ws.cell(row=row, column=3, value=percentage).border = self.thin_border
            row += 1
    
    def _create_keywords_sheet(self, ws, report_data: Dict[str, Any]):
        """创建关键词表"""
        # 表头
        headers = ['序号', '关键词', '权重']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        # 数据
        keywords = report_data.get('keywords', [])
        
        row = 2
        for i, kw in enumerate(keywords[:50], 1):  # 最多显示 50 个
            if isinstance(kw, dict):
                ws.cell(row=row, column=1, value=i).border = self.thin_border
                ws.cell(row=row, column=2, value=kw.get('word', '')).border = self.thin_border
                ws.cell(row=row, column=3, value=kw.get('weight', 1.0)).border = self.thin_border
            else:
                ws.cell(row=row, column=1, value=i).border = self.thin_border
                ws.cell(row=row, column=2, value=str(kw)).border = self.thin_border
                ws.cell(row=row, column=3, value=1.0).border = self.thin_border
            row += 1
    
    def _create_results_sheet(self, ws, report_data: Dict[str, Any]):
        """创建详细结果表"""
        # 表头
        headers = ['序号', '品牌', '问题', '模型', '状态', '创建时间']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        # 数据
        results = report_data.get('results', [])
        
        row = 2
        for i, result in enumerate(results, 1):
            ws.cell(row=row, column=1, value=i).border = self.thin_border
            ws.cell(row=row, column=2, value=result.get('brand', 'N/A')).border = self.thin_border
            ws.cell(row=row, column=3, value=result.get('question', 'N/A')).border = self.thin_border
            ws.cell(row=row, column=4, value=result.get('model', 'N/A')).border = self.thin_border
            ws.cell(row=row, column=5, value=result.get('status', 'N/A')).border = self.thin_border
            
            created_at = result.get('created_at', 'N/A')
            if created_at and created_at != 'N/A':
                try:
                    created_at = created_at[:19]  # 截取日期部分
                except:
                    pass
            ws.cell(row=row, column=6, value=created_at).border = self.thin_border
            
            row += 1
    
    def _auto_adjust_column_width(self, wb):
        """自动调整列宽"""
        for ws in wb.worksheets:
            column_widths = {}
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        col = cell.column
                        width = len(str(cell.value)) + 2
                        column_widths[col] = max(column_widths.get(col, 0), width)
            
            for col, width in column_widths.items():
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].width = min(width, 50)  # 最大 50
    
    def export_to_html(self, report_data: Dict[str, Any]) -> str:
        """
        导出报告为 HTML 格式（用于转长图）
        
        参数:
            report_data: 报告数据
            
        返回:
            str: HTML 字符串
        """
        try:
            html = f"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>AI 品牌诊断报告</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif; background: #f5f5f5; padding: 20px; }}
        .container {{ max-width: 800px; margin: 0 auto; background: white; border-radius: 12px; overflow: hidden; box-shadow: 0 2px 12px rgba(0,0,0,0.1); }}
        .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; text-align: center; }}
        .header h1 {{ font-size: 24px; margin-bottom: 10px; }}
        .header p {{ opacity: 0.9; font-size: 14px; }}
        .section {{ padding: 20px; border-bottom: 1px solid #eee; }}
        .section-title {{ font-size: 18px; font-weight: bold; color: #333; margin-bottom: 15px; display: flex; align-items: center; }}
        .section-title::before {{ content: ''; width: 4px; height: 18px; background: #667eea; margin-right: 10px; border-radius: 2px; }}
        .info-grid {{ display: grid; grid-template-columns: repeat(2, 1fr); gap: 15px; }}
        .info-item {{ background: #f8f9fa; padding: 12px; border-radius: 8px; }}
        .info-label {{ font-size: 12px; color: #666; margin-bottom: 5px; }}
        .info-value {{ font-size: 14px; color: #333; font-weight: 500; }}
        .brand-list {{ display: flex; flex-wrap: wrap; gap: 8px; }}
        .brand-tag {{ background: #e3f2fd; color: #1976d2; padding: 6px 12px; border-radius: 16px; font-size: 13px; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
        th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #eee; }}
        th {{ background: #f8f9fa; font-weight: 600; color: #333; font-size: 13px; }}
        td {{ font-size: 14px; color: #555; }}
        .status-completed {{ color: #28a745; }}
        .status-processing {{ color: #ffc107; }}
        .status-failed {{ color: #dc3545; }}
        .footer {{ text-align: center; padding: 20px; color: #999; font-size: 12px; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🎯 AI 品牌诊断报告</h1>
            <p>生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>
        
        <div class="section">
            <div class="section-title">报告概况</div>
            <div class="info-grid">
                <div class="info-item">
                    <div class="info-label">报告 ID</div>
                    <div class="info-value">{report_data.get('execution_id', 'N/A')[:8]}...</div>
                </div>
                <div class="info-item">
                    <div class="info-label">品牌名称</div>
                    <div class="info-value">{report_data.get('brand_name', 'N/A')}</div>
                </div>
                <div class="info-item">
                    <div class="info-label">诊断状态</div>
                    <div class="info-value {'status-' + report_data.get('status', 'N/A')}">{report_data.get('status', 'N/A')}</div>
                </div>
                <div class="info-item">
                    <div class="info-label">结果数量</div>
                    <div class="info-value">{len(report_data.get('results', []))}</div>
                </div>
            </div>
        </div>
        
        <div class="section">
            <div class="section-title">品牌分布</div>
            <table>
                <thead>
                    <tr>
                        <th>品牌名称</th>
                        <th>提及次数</th>
                        <th>占比</th>
                    </tr>
                </thead>
                <tbody>
"""
            
            # 品牌分布数据
            brand_dist = report_data.get('brandDistribution', {})
            data = brand_dist.get('data', {})
            total_count = brand_dist.get('totalCount', 0)
            
            for brand, count in (data or {}).items():
                percentage = f'{(count / total_count * 100):.1f}%' if total_count > 0 else '0%'
                html += f"""
                    <tr>
                        <td>{brand}</td>
                        <td>{count}</td>
                        <td>{percentage}</td>
                    </tr>
"""
            
            html += """
                </tbody>
            </table>
        </div>
        
        <div class="section">
            <div class="section-title">关键词云</div>
            <div class="brand-list">
"""
            
            # 关键词
            keywords = report_data.get('keywords', [])[:20]
            for kw in keywords:
                word = kw.get('word', str(kw)) if isinstance(kw, dict) else str(kw)
                html += f'<span class="brand-tag">{word}</span>'
            
            html += """
            </div>
        </div>
        
        <div class="footer">
            <p>本报告由 AI 品牌诊断系统自动生成</p>
            <p>© 2026 进化湾 GEO</p>
        </div>
    </div>
</body>
</html>
"""
            
            api_logger.info(f"[ExportService] ✅ HTML 导出成功")
            
            return html
            
        except Exception as e:
            api_logger.error(f"[ExportService] ❌ HTML 导出失败：{e}", exc_info=True)
            raise


# 全局单例
_export_service: Optional[ReportExportService] = None


def get_export_service() -> ReportExportService:
    """获取导出服务单例"""
    global _export_service
    if _export_service is None:
        _export_service = ReportExportService()
    return _export_service
